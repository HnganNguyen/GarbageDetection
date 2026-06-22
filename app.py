
import os
import json
import logging
import base64
import tempfile
from datetime import datetime
import numpy as np
import requests
import tensorflow as tf
from PIL import Image
from flask import Flask, request, render_template, url_for, redirect, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# CẤU HÌNH MODEL 3 NHÓM RÁC (Luật BVMT 2020 - Điều 75)
IMG_SIZE   = 260
MODEL_PATH = "waste_classifier_3groups_v2.keras"
# Nếu muốn dùng checkpoint tốt nhất, đổi sang:
# MODEL_PATH = "best_CustomCNN_3groups_v2.keras"

# 0 -> khac, 1 -> tai_che, 2 -> thuc_pham
IDX_TO_GROUP = {
    0: "khac",
    1: "tai_che",
    2: "thuc_pham",
}

# Tên hiển thị tiếng Việt
GROUP_VN_NAMES = {
    "tai_che"  : "Tái chế",
    "thuc_pham": "Thực phẩm",
    "khac"     : "Khác",
}

# Thông tin chi tiết cho 3 nhóm
GROUP_INFO = {
    "tai_che": {
        "vi_name"             : "Tái chế",
        "recyclability"       : "Có thể tái chế",
        "description"         : ("Chất thải rắn có khả năng tái sử dụng, tái chế: "
                                 "giấy, bìa các-tông, nhựa, kim loại, thủy tinh, "
                                 "vải, giày dép. Cần làm sạch trước khi bỏ vào thùng tái chế."),
        "environmental_impact": ("Nếu phân loại đúng, các vật liệu này được đưa vào "
                                 "vòng tuần hoàn tái chế, tiết kiệm tài nguyên thiên nhiên "
                                 "và giảm phát thải khí nhà kính."),
        "tips"                : ("Rửa sạch chai lọ, ép dẹp lon/hộp, gấp gọn bìa các-tông. "
                                 "Không lẫn rác ướt hoặc dầu mỡ vào thùng tái chế."),
        "recycle_centers"     : [
            "Trung tâm thu mua phế liệu địa phương",
            "Điểm thu gom của Veolia / Hợp tác xã Môi trường",
            "Chương trình đổi rác lấy quà của các siêu thị lớn",
        ],
    },
    "thuc_pham": {
        "vi_name"             : "Thực phẩm",
        "recyclability"       : "Hữu cơ",
        "description"         : ("Chất thải thực phẩm: thức ăn thừa, rau củ quả hỏng, "
                                 "vỏ trái cây, bã trà/cà phê, xương cá. Loại rác này "
                                 "phân hủy sinh học và có thể làm phân compost."),
        "environmental_impact": ("Khi chôn lấp lẫn với rác khác, thực phẩm phân hủy yếm "
                                 "khí sinh ra khí mê-tan — gấp 25 lần CO₂ về hiệu ứng nhà kính. "
                                 "Phân loại riêng giúp giảm phát thải đáng kể."),
        "tips"                : ("Để ráo nước trước khi bỏ vào thùng. Có thể tự ủ compost "
                                 "tại nhà cho vườn cây. Không bỏ lẫn túi nylon, đũa tre."),
        "recycle_centers"     : [
            "Điểm thu gom rác hữu cơ của phường/xã",
            "Trang trại đô thị nhận rác hữu cơ làm phân",
        ],
    },
    "khac": {
        "vi_name"             : "Khác",
        "recyclability"       : "Chất thải sinh hoạt khác",
        "description"         : ("Chất thải rắn sinh hoạt khác không thuộc 2 nhóm trên: "
                                 "pin, đồ điện tử hỏng, rác linh tinh khó tái chế. "
                                 "Lưu ý: pin và đồ điện tử là chất thải NGUY HẠI, cần "
                                 "thu gom riêng theo điểm tập kết chuyên dụng."),
        "environmental_impact": ("Pin chứa kim loại nặng gây ô nhiễm đất và nguồn nước "
                                 "nếu chôn lấp tùy tiện. Cần xử lý đúng quy trình chất thải nguy hại."),
        "tips"                : ("Pin/ắc-quy: mang đến điểm thu gom của các siêu thị "
                                 "hoặc điểm thu gom rác nguy hại tại địa phương. "
                                 "Tuyệt đối không vứt chung với rác sinh hoạt."),
        "recycle_centers"     : [
            "Việt Nam Tái Chế - vietnamrecycles.com",
            "Điểm thu gom pin tại các siêu thị lớn",
            "Công ty môi trường đô thị địa phương",
        ],
    },
}


#  load model .keras
@tf.keras.utils.register_keras_serializable(package="Custom")
class SimAMBlock(tf.keras.layers.Layer):
    """Simple, parameter-free 3D attention."""
    def __init__(self, e_lambda=1e-4, **kwargs):
        super().__init__(**kwargs)
        self.e_lambda = e_lambda

    def call(self, x):
        n    = tf.cast(tf.shape(x)[1] * tf.shape(x)[2], tf.float32)
        mean = tf.reduce_mean(x, axis=[1, 2], keepdims=True)
        d    = tf.square(x - mean)
        var  = tf.reduce_sum(d, axis=[1, 2], keepdims=True) / (n - 1)
        attn = tf.sigmoid(d / (4.0 * (var + self.e_lambda)) + 0.5)
        return x * attn

    def get_config(self):
        cfg = super().get_config()
        cfg["e_lambda"] = self.e_lambda
        return cfg


def get_group_info(group_key):
    return GROUP_INFO.get(group_key, {
        "vi_name"             : "Không xác định",
        "recyclability"       : "Không xác định",
        "description"         : "Không có thông tin",
        "environmental_impact": "Không có thông tin",
        "tips"                : "Không có thông tin",
        "recycle_centers"     : [],
    })

# Setup logging
logging.basicConfig(level=logging.INFO)

app = Flask(__name__)
app.secret_key = 'your_secret_key'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = './uploads'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///waste_classification.db'
app.secret_key = 'your_secret_key'
db = SQLAlchemy(app)

# Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), nullable=False, unique=True)
    password = db.Column(db.String(255))
    points = db.Column(db.Integer, default=0)

class CustomerData(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    email = db.Column(db.String(150), nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    address = db.Column(db.String(255), nullable=True)
    user = db.relationship("User", backref="customer_data")

class UserActivity(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action = db.Column(db.String(255), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    user = db.relationship("User", backref="activities")

# Ensure the database is created
with app.app_context():
    db.create_all()

# Load Model

def load_model(model_path):
    try:
        model = tf.keras.models.load_model(
            model_path,
            custom_objects={"SimAMBlock": SimAMBlock},
            compile=False,
        )
        logging.info(f"Model loaded successfully from {model_path}")
        logging.info(f"Input shape : {model.input_shape}")
        logging.info(f"Output shape: {model.output_shape}")
        return model
    except Exception as e:
        logging.error(f"Error loading model: {e}")
        return None


model = load_model(MODEL_PATH)


def classify_image(image_path):
    """
    Trả về tuple (group_key, confidence) hoặc (None, 0.0) nếu lỗi.
    group_key: 'tai_che' | 'thuc_pham' | 'khac'
    confidence: độ tin cậy từ 0 đến 1
    """
    try:
        if model is None:
            logging.error("Model not loaded.")
            return None, 0.0

        image = Image.open(image_path).convert("RGB")
        image = image.resize((IMG_SIZE, IMG_SIZE))
        arr   = np.array(image, dtype=np.float32) / 255.0 
        arr   = np.expand_dims(arr, axis=0)

        preds      = model.predict(arr, verbose=0)[0]
        class_idx  = int(np.argmax(preds))
        confidence = float(preds[class_idx])
        group_key  = IDX_TO_GROUP.get(class_idx)

        logging.info(f"Prediction: {group_key} (confidence={confidence:.4f})")
        return group_key, confidence
    except Exception as e:
        logging.error(f"Error during classification: {e}")
        return None, 0.0


from werkzeug.security import generate_password_hash

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username'].strip()
        email = request.form['email'].strip()
        phone = request.form['phone'].strip()
        address = request.form['address'].strip()
        password = request.form['password']
        confirm_password = request.form['confirm_password']

        if not username:
            flash("Tên người dùng không thể để trống!", "danger")
            return redirect('/register')

        if User.query.filter_by(username=username).first():
            flash("Tên người dùng đã tồn tại!", "danger")
            return redirect('/register')

        # kiểm tra xác nhận mật khẩu
        if password != confirm_password:
            flash("Mật khẩu xác nhận không đúng!", "danger")
            return redirect('/register')

        # mã hóa password
        hashed_password = generate_password_hash(password)

        # tạo user
        new_user = User(username=username, password=hashed_password)
        db.session.add(new_user)
        db.session.commit()

        # lưu thông tin khách hàng
        customer_data = CustomerData(
            user_id=new_user.id,
            email=email,
            phone=phone,
            address=address
        )

        db.session.add(customer_data)
        db.session.commit()

        flash("Đăng ký thành công! Bạn có thể đăng nhập ngay bây giờ.", "success")
        return redirect('/login')

    return render_template('register.html')

from werkzeug.security import check_password_hash

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if not user:
            flash("Username không tồn tại!", "danger")
            return redirect('/login')
        if not user.password:
            flash("Tài khoản chưa có mật khẩu!", "danger")
            return redirect('/login')
        # kiểm tra mật khẩu đã mã hóa
        if not check_password_hash(user.password, password):

            flash("Sai mật khẩu!", "danger")
            return redirect('/login')
        session['user_id'] = user.id

        flash("Đăng nhập thành công!", "success")
        
        return redirect('/home')

    return render_template('login.html')

@app.route('/logout')
def logout():
    if 'user_id' in session:
        user_id = session.pop('user_id')
        activity = UserActivity(user_id=user_id, action="Logged out")
        db.session.add(activity)
        db.session.commit()
    flash("Successfully logged out!", "info")
    return redirect('/login')

from sqlalchemy import func, desc

@app.route('/change_password', methods=['POST'])
def change_password():
    if 'user_id' not in session:
        return redirect('/login')
    user = User.query.get(session['user_id'])
    old_password = request.form['old_password']
    new_password = request.form['new_password']
    confirm_password = request.form['confirm_password']

    # kiểm tra mật khẩu cũ
    if not check_password_hash(user.password, old_password):

        flash("Mật khẩu hiện tại không đúng", "danger")

        return redirect('/home')
    # kiểm tra xác nhận
    if new_password != confirm_password:

        flash("Mật khẩu xác nhận không khớp", "danger")

        return redirect('/home')
    # cập nhật mật khẩu
    user.password = generate_password_hash(new_password)

    db.session.commit()

    # lưu log
    activity = UserActivity(

        user_id=user.id,

        action="Đổi mật khẩu"

    )

    db.session.add(activity)
    db.session.commit()

    flash("Đổi mật khẩu thành công!", "success")
    return redirect('/home')

@app.route('/home')
def home():

    if 'user_id' not in session:
        return redirect('/login')
    user = db.session.get(User, session['user_id'])
    is_admin = user.username == "admin"
    username = user.username
    total_users = User.query.count()

    total_classify = db.session.query(
        func.count(UserActivity.id)
    ).filter(
        UserActivity.action.like("Phân loại:%")
    ).scalar()

    top_user_data = User.query.order_by(User.points.desc()).first()
    top_user = top_user_data.username if top_user_data else "None"

    # 🔥 Thống kê rác từ bảng activity

    waste_stats = db.session.query(
    UserActivity.action,
    func.count(UserActivity.id)
    ).filter(
    UserActivity.action.like("Phân loại:%")
    ).group_by(UserActivity.action).all()

    waste_labels = []
    waste_counts = []

    for action, count in waste_stats:
        waste = action.split(":")[1].strip()
        waste_labels.append(waste)
        waste_counts.append(count)
    # Top waste
    if waste_counts:
        max_index = waste_counts.index(max(waste_counts))
        top_waste = waste_labels[max_index]
    else:
        top_waste = "None"

    # Recent activity
    recent_activity = db.session.query(
    User.username,
    UserActivity.action,
    UserActivity.timestamp
    ).join(
    User, User.id == UserActivity.user_id
    ).order_by(
    UserActivity.timestamp.desc()
    ).limit(5).all()
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    return render_template(
           "home.html",
     username=user.username,
     is_admin=is_admin,
     now=now,
     total_users=total_users,
     total_classify=total_classify,
     top_waste=top_waste,
     top_user=top_user,
     recent_activity=recent_activity,
     waste_labels=waste_labels,
     waste_counts=waste_counts
    )


@app.route('/export_excel')
def export_excel():
    # phải đăng nhập
    if 'user_id' not in session:
        return redirect('/login')
    user = User.query.get(session['user_id'])
    # chỉ admin mới được xuất
    if user.username != "admin":
        return "Bạn không có quyền xuất file!", 403
    # 1. dữ liệu chi tiết
    details = db.session.query(
        User.username,
        UserActivity.action,
        UserActivity.timestamp
    ).join(
        User, User.id == UserActivity.user_id
    ).filter(
        UserActivity.action.like("Phân loại:%")
    ).order_by(
        UserActivity.timestamp.desc()
    ).all()

    # 2. thống kê loại rác
    waste_stats = db.session.query(
        UserActivity.action,
        func.count(UserActivity.id)
    ).filter(
        UserActivity.action.like("Phân loại:%")
    ).group_by(
        UserActivity.action
    ).all()

    # 3. top user
    top_users = db.session.query(
        User.username,
        func.count(UserActivity.id).label("total")
    ).join(
        UserActivity, User.id == UserActivity.user_id
    ).filter(
        UserActivity.action.like("Phân loại:%")
    ).group_by(
        User.username
    ).order_by(
        func.count(UserActivity.id).desc()
    ).all()

    wb = Workbook()
    # style chung
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    center_align = Alignment(horizontal="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Sheet 1: chi tiết
    ws1 = wb.active
    ws1.title = "ChiTiet"

    headers = ["Username", "Loại rác", "Thời gian"]

    ws1.append(headers)

    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for username, action, time in details:

        waste = action.split(":")[1].strip()

        ws1.append([
            username,
            waste,
            time.strftime("%d/%m/%Y %H:%M")
        ])

    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 25
    # Sheet 2: thống kê rác

    ws2 = wb.create_sheet("ThongKe_Rac")

    ws2.append(["Loại rác", "Số lần"])

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for action, total in waste_stats:

        waste = action.split(":")[1].strip()

        ws2.append([
            waste,
            total
        ])

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    # Sheet 3: top user
    ws3 = wb.create_sheet("Top_User")

    ws3.append(["Username", "Số lần phân loại"])

    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for username, total in top_users:

        ws3.append([
            username,
            total
        ])

    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 20
    # lưu file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

    wb.save(tmp.name)

    tmp.close()

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="ThongKe_PhanLoaiRac.xlsx"
    )

@app.route('/', methods=['GET', 'POST'])
def upload_image():
    if 'user_id' not in session:
        flash("Vui lòng đăng nhập để sử dụng tính năng này!", "warning")
        return redirect('/login')

    user = db.session.get(User, session['user_id'])
    is_admin = user.username == "admin"

    if request.method == 'POST':
        # Xử lý ảnh chụp từ camera (Base64)
        captured_image = request.form.get('captured_image')
        if captured_image:
            print("Ảnh từ camera nhận được!")
            try:
                # Decode Base64 và lưu file
                image_data = base64.b64decode(captured_image.split(',')[1])
                filename = f"captured_{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                with open(filepath, "wb") as f:
                    f.write(image_data)
                print(f"Ảnh đã được lưu tại: {filepath}")

                # Chuyển file sang static/uploads để sử dụng trên giao diện web
                static_path = os.path.join('static', 'uploads', filename)
                os.makedirs(os.path.dirname(static_path), exist_ok=True)
                if not os.path.exists(static_path):
                    os.rename(filepath, static_path)

                # Phân loại ảnh
                result_key, confidence = classify_image(static_path)
                print(f"Kết quả phân loại: {result_key}, confidence={confidence:.4f}")

                if result_key:
                    # Cộng điểm
                    user.points += 10
                    db.session.commit()

                    # Lấy thông tin nhóm
                    info = get_group_info(result_key)
                    vi_name = info["vi_name"]

                    # Lưu hoạt động người dùng
                    activity = UserActivity(
                        user_id=user.id,
                        action=f"Phân loại: {vi_name}"
                    )
                    db.session.add(activity)
                    db.session.commit()

                    return render_template(
                        'result.html',
                        result=vi_name,
                        description=info["description"],
                        environmental_impact=info["environmental_impact"],
                        tips=info["tips"],
                        recycle_centers=info["recycle_centers"],
                        image_url=url_for('static', filename=f'uploads/{filename}'),
                        recyclability=info["recyclability"],
                        recyclability_description=info["description"],
                        confidence=round(confidence * 100, 2),
                        group_key=result_key,
                        is_admin=is_admin
                    )
                else:
                    flash("Xảy ra lỗi trong quá trình phân loại. Vui lòng thử lại!", "danger")
            except Exception as e:
                flash(f"Lỗi xử lý ảnh từ camera: {e}", "danger")
                return redirect('/')

        # Xử lý ảnh tải lên từ file
        file = request.files.get('file')
        if file and file.filename != '':
            # Lưu file vào thư mục uploads
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            # Chuyển file sang static/uploads để sử dụng trên giao diện web
            static_path = os.path.join('static', 'uploads', filename)
            os.makedirs(os.path.dirname(static_path), exist_ok=True)
            if not os.path.exists(static_path):
                os.rename(filepath, static_path)

            # Phân loại ảnh
            result_key, confidence = classify_image(static_path)
            print(f"Kết quả phân loại: {result_key}, confidence={confidence:.4f}")

            if result_key:
                # Cộng điểm cho người dùng
                user.points += 10
                db.session.commit()

                # Lấy thông tin nhóm
                info = get_group_info(result_key)
                vi_name = info["vi_name"]

                # Lưu hoạt động người dùng
                activity = UserActivity(
                    user_id=user.id,
                    action=f"Phân loại: {vi_name}"
                )
                db.session.add(activity)
                db.session.commit()

                # Trả về kết quả
                return render_template(
                    'result.html',
                    result=vi_name,
                    description=info["description"],
                    environmental_impact=info["environmental_impact"],
                    tips=info["tips"],
                    recycle_centers=info["recycle_centers"],
                    image_url=url_for('static', filename=f'uploads/{filename}'),
                    recyclability=info["recyclability"],
                    recyclability_description=info["description"],
                    confidence=round(confidence * 100, 2),
                    group_key=result_key,
                    is_admin=is_admin
                )
            else:
                flash("Xảy ra lỗi trong quá trình phân loại. Vui lòng thử lại!", "danger")
        else:
            flash("Chưa chọn tệp hoặc chụp ảnh!", "danger")
    # Hiển thị giao diện upload
    return render_template('upload.html', is_admin=is_admin)

@app.route('/leaderboard')
def leaderboard():
    users = User.query.order_by(User.points.desc()).all()
    is_admin = 'user_id' in session and User.query.get(session['user_id']).username == "admin"
    return render_template('leaderboard.html', users=users, is_admin=is_admin)

@app.route('/history')
def history():
    if 'user_id' not in session:
        flash("Please log in to view activity history!", "warning")
        return redirect('/login')

    user_id = session['user_id']
    activities = UserActivity.query.filter_by(user_id=user_id).order_by(UserActivity.timestamp.desc()).all()
    is_admin = User.query.get(user_id).username == "admin"
    return render_template('history.html', activities=activities, is_admin=is_admin)

@app.route('/admin')
def admin_dashboard():
    if 'user_id' not in session:
        return redirect('/login')

    admin_user = User.query.get(session['user_id'])
    if admin_user.username != "admin":
        return "You do not have access!", 403

    users = User.query.all()
    customer_data = CustomerData.query.all()
    activities = UserActivity.query.order_by(UserActivity.timestamp.desc()).all()

    return render_template('admin.html', users=users, customer_data=customer_data, activities=activities, is_admin=True)

# ===== Khởi động ứng dụng =====

if __name__ == "__main__":
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs('./static/uploads', exist_ok=True)

    if model is None:
        print(f"❌ Không load được model! Kiểm tra file {MODEL_PATH}")
        exit(1)

    print("✅ Model ready!")

    app.run(
        host="0.0.0.0",
        port=8080,
        debug=False
    )








  
